import asyncio
import io
import logging
import math
import re
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from aiogram import Bot, Dispatcher, F
from aiogram.filters import Command, CommandStart
from aiogram.types import (
    Message, CallbackQuery, BufferedInputFile,
    InlineKeyboardMarkup, InlineKeyboardButton,
)

from config import TELEGRAM_BOT_TOKEN
from scraper import SearchManager, detect_platforms, clean_query, PLATFORM_LABELS
from ai_agent import GeminiAgent
from scraper_web import WebScraper
from database import (
    ensure_database_exists, init_db, close_pool,
    save_schedule, delete_schedule, get_all_schedules,
    get_user_settings, save_user_settings,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger(__name__)

bot = Bot(token=TELEGRAM_BOT_TOKEN)
dp = Dispatcher()
search_manager = SearchManager()
agent = GeminiAgent()

# Per-user in-memory state (non-persistent)
user_tasks: dict[int, asyncio.Task] = {}
user_context: dict[int, bool] = {}
scheduled_tasks: dict[int, asyncio.Task] = {}

FILTER_QUALIFIERS = {"лише", "тільки", "лиш", "саме"}

GROUP_LABELS = {
    "Adult":      "Adult — для дорослих котів",
    "Sterilized": "Sterilized — для стерилізованих",
    "Kitten":     "Kitten — для кошенят",
    "Senior":     "Senior — для літніх котів (7+)",
    "Hairball":   "Hairball Control — виведення шерсті",
    "Sensitive":  "Sensitive — чутливе травлення",
    "Indoor":     "Indoor — для домашніх котів",
    "Інше":       "Інше",
}
GROUP_ORDER = ["Adult", "Sterilized", "Kitten", "Senior", "Hairball", "Sensitive", "Indoor", "Інше"]


# ------------------------------------------------------------------ #
#  Keyboards                                                           #
# ------------------------------------------------------------------ #

def settings_keyboard(platforms: list[str]) -> InlineKeyboardMarkup:
    """Inline keyboard: одиночний вибір платформи (лише одна активна)."""
    row = [
        InlineKeyboardButton(
            text=("✅ " if p in platforms else "◻️ ") + label,
            callback_data=f"platform:{p}",
        )
        for p, label in [("prom", "Prom"), ("rozetka", "Rozetka")]
    ]
    return InlineKeyboardMarkup(inline_keyboard=[row])


# ------------------------------------------------------------------ #
#  Output formatters                                                   #
# ------------------------------------------------------------------ #

def format_chat(query: str, platform: str, products: list[dict], ai_analysis: str) -> str:
    label = PLATFORM_LABELS.get(platform, platform)
    lines = [f'🔍 Результати пошуку: "{query}" [{label}]\n']
    if not products:
        lines.append("😔 Товари не знайдено. Спробуйте інший запит або платформу.")
        return "\n".join(lines)

    lines.append(f"📦 Знайдено {len(products)} товарів:\n")
    for i, p in enumerate(products, 1):
        lines.append(f"{i}. {p['name']}")
        lines.append(f"   💰 {p['price']}")
        if p.get("city"):
            lines.append(f"   📍 {p['city']}")
        if p.get("url"):
            lines.append(f"   🔗 {p['url']}")
        lines.append("")

    lines.append("📊 AI Аналіз:")
    lines.append(ai_analysis)
    return "\n".join(lines)


def _parse_price_int(price_str: str) -> int | None:
    """'275,00 грн' → 275,  '2 864.20₴' → 2864,  '1 500 грн' → 1500,  'Ціна не вказана' → None"""
    s = re.sub(r"[^\d\s,.]", "", price_str).strip()  # залишаємо цифри, пробіли, кому, крапку
    s = re.sub(r"[,.](\d{1,2})\s*$", "", s).strip()  # видаляємо десяткову частину: ,00 або .20
    s = re.sub(r"[\s,.]", "", s)                      # прибираємо роздільники тисяч
    return int(s) if s.isdigit() else None


def build_excel(query: str, platform: str, products: list[dict], ai_analysis: str) -> bytes:
    """Single sheet: title → column headers → [grouped or flat] products → AI analysis."""
    from collections import defaultdict

    def _price_key(p: dict) -> int:
        n = _parse_price_int(p.get("price", ""))
        return n if n is not None else 10 ** 9

    has_groups = any(p.get("group") for p in products)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Результати"

    show_seller = platform not in ("olx", "rozetka")
    show_city   = platform != "rozetka"
    if show_seller and show_city:
        col_widths = [5, 60, 28, 28, 22, 28, 16]
        headers    = ["№", "Назва", "Ціна", "Продавець", "Місто", "Посилання", "Платформа"]
        last_col   = "G"
        n_cols     = 7
    elif show_city:
        col_widths = [5, 65, 28, 25, 30, 16]
        headers    = ["№", "Назва", "Ціна", "Місто", "Посилання", "Платформа"]
        last_col   = "F"
        n_cols     = 6
    else:
        col_widths = [5, 70, 30, 32, 16]
        headers    = ["№", "Назва", "Ціна", "Посилання", "Платформа"]
        last_col   = "E"
        n_cols     = 5

    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = w

    # --- Styles ---
    blue_fill    = PatternFill("solid", fgColor="1F4E79")
    group_fill   = PatternFill("solid", fgColor="2E75B6")
    summary_fill = PatternFill("solid", fgColor="FFF2CC")
    alt_fill     = PatternFill("solid", fgColor="D6E4F0")
    ai_fill      = PatternFill("solid", fgColor="E2EFDA")
    white_font   = Font(color="FFFFFF", bold=True)
    bold_font    = Font(bold=True)
    link_font    = Font(color="0563C1", underline="single")
    center       = Alignment(horizontal="center", vertical="center")
    wrap         = Alignment(wrap_text=True, vertical="top")

    # --- Row 1: title ---
    ws.merge_cells(f"A1:{last_col}1")
    tc = ws["A1"]
    tc.value = (
        f'Пошук: "{query}"  |  '
        f'Платформа: {PLATFORM_LABELS.get(platform, platform)}  |  '
        f'Дата: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
    )
    tc.font = Font(bold=True, size=12, color="FFFFFF")
    tc.fill = blue_fill
    tc.alignment = center
    ws.row_dimensions[1].height = 22

    # --- Row 2: column headers ---
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = white_font
        cell.fill = blue_fill
        cell.alignment = center

    # --- Helper: render one product row ---
    def render_row(row: int, num: int, p: dict, is_alt: bool) -> None:
        col = 1
        ws.cell(row=row, column=col, value=num).alignment = center;          col += 1
        ws.cell(row=row, column=col, value=p.get("name", "")).alignment = wrap; col += 1
        ws.cell(row=row, column=col, value=p.get("price", "")).alignment = center; col += 1
        if show_seller:
            seller = p.get("seller", "")
            seller = seller[:30] + "…" if len(seller) > 30 else seller
            ws.cell(row=row, column=col, value=seller).alignment = wrap; col += 1
        if show_city:
            ws.cell(row=row, column=col, value=p.get("city", "")).alignment = wrap;  col += 1
        url = p.get("url", "")
        lc = ws.cell(row=row, column=col, value="Відкрити →" if url else "")
        if url:
            lc.hyperlink = url
            lc.font = link_font
        lc.alignment = center; col += 1
        ws.cell(row=row, column=col,
                value=PLATFORM_LABELS.get(p.get("platform", platform), platform)).alignment = center
        if is_alt:
            for c in range(1, n_cols + 1):
                ws.cell(row=row, column=c).fill = alt_fill

    current_row = 3

    if has_groups:
        # --- GROUPED rendering ---
        grouped: dict = defaultdict(list)
        for p in products:
            grouped[p.get("group", "Інше")].append(p)

        # Спочатку відомі групи в порядку GROUP_ORDER, потім динамічні (смак+вага) за алфавітом
        known = [k for k in GROUP_ORDER if k in grouped]
        dynamic = sorted(k for k in grouped if k not in GROUP_ORDER)
        all_keys = known + dynamic

        flat_items = []  # товари з груп < 3 — виводимо плоско в кінці

        row_counter = 1
        for group_key in all_keys:
            gp = sorted(grouped[group_key], key=_price_key)
            label = GROUP_LABELS.get(group_key, group_key)

            if len(gp) < 3:
                # Мало товарів — відкладаємо у плоский список
                flat_items.extend(gp)
                continue

            # Group header
            ws.merge_cells(f"A{current_row}:{last_col}{current_row}")
            gh = ws.cell(row=current_row, column=1,
                         value=f"▶  {label}  ({len(gp)} товарів)")
            gh.font = Font(bold=True, size=11, color="FFFFFF")
            gh.fill = group_fill
            gh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[current_row].height = 20
            current_row += 1

            # Products
            for i, p in enumerate(gp, 1):
                render_row(current_row, i, p, i % 2 == 0)
                current_row += 1

            # Summary row
            nums = [n for p in gp if (n := _parse_price_int(p.get("price", ""))) is not None]
            if nums:
                summary = (
                    f"  Мін: {min(nums):,} грн   |   "
                    f"Макс: {max(nums):,} грн   |   "
                    f"Середня: {sum(nums)/len(nums):,.0f} грн"
                ).replace(",", " ")
            else:
                summary = "  Ціни недоступні"
            ws.merge_cells(f"A{current_row}:{last_col}{current_row}")
            sc = ws.cell(row=current_row, column=1, value=f"📊{summary}")
            sc.font = Font(bold=True, size=10)
            sc.fill = summary_fill
            sc.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[current_row].height = 18
            current_row += 1

            # Separator
            ws.row_dimensions[current_row].height = 8
            current_row += 1

        # Плоский хвіст — товари з груп < 3
        for i, p in enumerate(sorted(flat_items, key=_price_key), 1):
            render_row(current_row, i, p, i % 2 == 0)
            current_row += 1

    else:
        # --- FLAT rendering (fallback) ---
        for i, p in enumerate(sorted(products, key=_price_key), 3):
            render_row(i, i - 2, p, i % 2 == 0)
        current_row = len(products) + 3

    # --- AI Analysis ---
    ws.row_dimensions[current_row].height = 8
    ai_start = current_row + 1
    ws.merge_cells(f"A{ai_start}:{last_col}{ai_start}")
    hc = ws.cell(row=ai_start, column=1, value="📊 AI Аналіз")
    hc.font = Font(bold=True, size=11, color="FFFFFF")
    hc.fill = blue_fill
    hc.alignment = center

    for offset, line in enumerate(ai_analysis.splitlines(), 1):
        row_idx = ai_start + offset
        ws.merge_cells(f"A{row_idx}:{last_col}{row_idx}")
        cell = ws.cell(row=row_idx, column=1, value=line)
        cell.fill = ai_fill
        cell.font = bold_font if line.strip().startswith("•") else Font()
        cell.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
        ws.row_dimensions[row_idx].height = max(1, math.ceil(len(line) / 120)) * 15

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ------------------------------------------------------------------ #
#  Helpers                                                             #
# ------------------------------------------------------------------ #

def _filter_by_price(products: list[dict], filters: dict) -> list[dict]:
    """Програмна фільтрація по ціні — до передачі в AI."""
    price_min = filters.get("price_min")
    price_max = filters.get("price_max")
    if price_min is None and price_max is None:
        return products
    result = []
    for p in products:
        price = _parse_price_int(p.get("price", ""))
        if price is None:
            result.append(p)  # ціна невідома — не відкидаємо
            continue
        if price_min is not None and price < price_min:
            continue
        if price_max is not None and price > price_max:
            continue
        result.append(p)
    return result


# ------------------------------------------------------------------ #
#  Core search function                                                #
# ------------------------------------------------------------------ #

async def _collect_excel_results(
    user_id: int, query: str, platform: str,
    filter_intent: str, filters: dict | None = None,
    target: int = 100, max_pages: int = 25,
) -> list[dict]:
    """2 браузери паралельно + пайплайн: поки Gemini фільтрує пару N,
    браузери вже тягнуть пару N+1.

    AI працює в двох режимах:
      - ШИРОКИЙ: немає уточнень → бере всі варіанти категорії
      - СУВОРИЙ: є filter_intent або filters → тільки точні збіги
    """
    collected: list[dict] = []
    seen: set[str] = set()

    async def fetch_pair(p: int) -> list[dict]:
        """Два браузери одночасно: сторінки p і p+1."""
        if p + 1 <= max_pages:
            a, b = await asyncio.gather(
                asyncio.to_thread(search_manager.search_page, query, platform, p),
                asyncio.to_thread(search_manager.search_page, query, platform, p + 1),
            )
            return a + b
        return await asyncio.to_thread(search_manager.search_page, query, platform, p)

    page = 1
    # Запускаємо першу пару одразу
    fetch_task: asyncio.Task = asyncio.create_task(fetch_pair(page))

    while page <= max_pages:
        raw = await fetch_task
        page += 2

        if not raw:
            break

        # Пайплайн: запускаємо наступну пару ДО початку фільтрації
        if page <= max_pages:
            fetch_task = asyncio.create_task(fetch_pair(page))
        else:
            fetch_task = None

        # Програмна фільтрація по ціні (швидко, до AI)
        raw = _filter_by_price(raw, filters or {})

        # --- LOG: що прийшло зі скрапера ---
        print(f"\n{'='*60}")
        print(f"[SCRAPER] Запит: '{query}' | Платформа: {platform}")
        print(f"[SCRAPER] Сторінки {page-2}-{page-1} | Знайдено: {len(raw)} товарів")
        for i, p in enumerate(raw, 1):
            print(f"  {i:>3}. {p.get('name','?')[:70]} | {p.get('price','?')}")
        print(f"{'='*60}")

        # Фільтрація поточної пари (Gemini працює поки браузери тягнуть наступну)
        batch = await asyncio.to_thread(
            agent.filter_products_by_intent,
            user_id, raw, query, filter_intent, filters or {},
        )

        # --- LOG: що залишив AI ---
        kept_names = {p.get('name') for p in batch}
        print(f"[AI FILTER] Залишив: {len(batch)} з {len(raw)}")
        for i, p in enumerate(raw, 1):
            name = p.get('name', '?')
            status = "✅ ЗАЛИШИВ" if name in kept_names else "❌ ВІДРІЗАВ"
            print(f"  {i:>3}. [{status}] {name[:70]} | {p.get('price','?')}")
        print(f"{'='*60}\n")

        for product in batch:
            uid = product.get("product_id") or product.get("url") or ""
            if uid and uid in seen:
                continue
            if uid:
                seen.add(uid)
            collected.append(product)

        if len(collected) >= target or fetch_task is None:
            if fetch_task and not fetch_task.done():
                fetch_task.cancel()
            break

    return collected[:target]


async def _search_one_platform(
    user_id: int, chat_id: int, query: str,
    platform: str, filter_intent: str = "",
) -> None:
    """Збирає та надсилає Excel для однієї платформи."""
    products = await _collect_excel_results(user_id, query, platform, filter_intent)
    products = await asyncio.to_thread(agent.group_products_by_subtype, user_id, products)
    ai_analysis = await asyncio.to_thread(agent.analyze_prices, user_id, query, products, platform)
    xlsx_bytes = await asyncio.to_thread(build_excel, query, platform, products, ai_analysis)
    filename = f"price_{query[:25].replace(' ', '_')}_{platform}.xlsx"
    await bot.send_document(
        chat_id,
        BufferedInputFile(xlsx_bytes, filename=filename),
        caption=f'📊 {PLATFORM_LABELS.get(platform, platform)}: "{query}"',
    )


async def do_search(
    user_id: int, chat_id: int, query: str,
    platforms: list[str], status_msg=None,
    filter_intent: str = "", filters: dict | None = None,
) -> None:
    """Паралельний пошук по всіх платформах — окремий Excel на кожну."""
    gather_tasks = [
        _collect_excel_results(user_id, query, p, filter_intent, filters)
        for p in platforms
    ]
    results_list: list[list[dict]] = await asyncio.gather(*gather_tasks)

    if status_msg:
        try:
            await status_msg.delete()
        except Exception:
            pass

    # Формуємо і надсилаємо Excel по черзі (щоб не флудити одночасно)
    for platform, products in zip(platforms, results_list):
        products = await asyncio.to_thread(agent.group_products_by_subtype, user_id, products)
        ai_analysis = await asyncio.to_thread(
            agent.analyze_prices, user_id, query, products, platform
        )
        xlsx_bytes = await asyncio.to_thread(build_excel, query, platform, products, ai_analysis)
        filename = f"price_{query[:25].replace(' ', '_')}_{platform}.xlsx"
        await bot.send_document(
            chat_id,
            BufferedInputFile(xlsx_bytes, filename=filename),
            caption=f'📊 {PLATFORM_LABELS.get(platform, platform)}: "{query}"',
        )

    user_context[user_id] = True


def _start_schedule_task(
    user_id: int, chat_id: int, interval: int,
    query: str, platforms: list[str],
) -> asyncio.Task:
    async def loop():
        while True:
            await do_search(user_id, chat_id, query, platforms)
            await asyncio.sleep(interval * 60)

    task = asyncio.create_task(loop())
    scheduled_tasks[user_id] = task
    return task


# ------------------------------------------------------------------ #
#  Startup / shutdown                                                  #
# ------------------------------------------------------------------ #

@dp.startup()
async def on_startup() -> None:
    await ensure_database_exists()
    await init_db()

    schedules = await get_all_schedules()
    for s in schedules:
        platforms = [p for p in s.get("platform", "prom").split(",") if p]
        _start_schedule_task(
            s["user_id"], s["chat_id"], s["interval_minutes"],
            s["query"], platforms,
        )
    if schedules:
        logger.info("Restored %d scheduled task(s) from DB", len(schedules))


@dp.shutdown()
async def on_shutdown() -> None:
    await close_pool()


# ------------------------------------------------------------------ #
#  Commands                                                            #
# ------------------------------------------------------------------ #

@dp.message(CommandStart())
async def cmd_start(message: Message) -> None:
    user_id = message.from_user.id
    agent.reset_chat(user_id)
    user_context.pop(user_id, None)
    await message.answer(
        "👋 Привіт! Я бот для пошуку цін.\n\n"
        "Напишіть що шукати, наприклад:\n"
        "Вологий корм Optimeal для котів з ягням та овочами в желе 85 г в пром ціна до 100 грн"
    )


@dp.message(Command("settings"))
async def cmd_settings(message: Message) -> None:
    settings = await get_user_settings(message.from_user.id)
    plat_labels = ", ".join(PLATFORM_LABELS.get(p, p) for p in settings["platforms"])
    await message.answer(
        f"⚙️ Поточні налаштування:\n"
        f"Платформи: {plat_labels}\n\n"
        f"Натисніть щоб увімкнути/вимкнути платформу:",
        reply_markup=settings_keyboard(settings["platforms"]),
    )


@dp.callback_query(F.data.startswith("platform:"))
async def on_platform_select(callback: CallbackQuery) -> None:
    user_id = callback.from_user.id
    platform = callback.data.split(":")[1]
    await save_user_settings(user_id, platforms=[platform])
    await callback.answer(f"✅ Платформа: {PLATFORM_LABELS.get(platform, platform)}")
    await callback.message.edit_reply_markup(
        reply_markup=settings_keyboard([platform])
    )



@dp.message(Command("reboot"))
async def cmd_reboot(message: Message) -> None:
    user_id = message.from_user.id
    task = user_tasks.pop(user_id, None)
    if task and not task.done():
        task.cancel()
    agent.reset_chat(user_id)
    user_context.pop(user_id, None)
    await message.answer("🔄 AI перезапущено. Контекст очищено!")


# ------------------------------------------------------------------ #
#  Message handlers                                                    #
# ------------------------------------------------------------------ #

@dp.message(F.text)
async def handle_text(message: Message) -> None:
    user_id = message.from_user.id
    text = message.text.strip()
    if not text:
        return

    # Ignore unknown slash-commands
    if text.startswith("/"):
        return

    settings = await get_user_settings(user_id)

    # AI класифікує намір користувача
    intent = await asyncio.to_thread(agent.classify_intent, user_id, text)
    action = intent.get("action", "chat")

    # ------------------------------------------------------------------ #
    # platform_info: яка платформа зараз
    # ------------------------------------------------------------------ #
    if action == "platform_info":
        label = PLATFORM_LABELS.get(settings["platforms"][0], settings["platforms"][0])
        await message.answer(f"📍 Зараз шукаємо на: {label}")
        return

    # ------------------------------------------------------------------ #
    # platform_switch: переключити платформу
    # ------------------------------------------------------------------ #
    if action == "platform_switch":
        platform = (intent.get("platforms") or [""])[0]
        if platform == "olx":
            await message.answer(
                "⛔ OLX наразі недоступний. Оберіть іншу платформу:",
                reply_markup=settings_keyboard(settings["platforms"]),
            )
        elif platform in ("prom", "rozetka"):
            await save_user_settings(user_id, platforms=[platform])
            label = PLATFORM_LABELS.get(platform, platform)
            await message.answer(f"✅ Платформу змінено на {label}")
        else:
            await message.answer(
                "❓ Не розпізнав платформу. Доступні: Prom, Rozetka",
                reply_markup=settings_keyboard(settings["platforms"]),
            )
        return

    # ------------------------------------------------------------------ #
    # cancel_search: скасувати активний пошук
    # ------------------------------------------------------------------ #
    if action == "cancel_search":
        task = user_tasks.pop(user_id, None)
        if task and not task.done():
            task.cancel()
            await message.answer("⛔ Пошук скасовано.")
        else:
            await message.answer("Активного пошуку немає.")
        return

    # ------------------------------------------------------------------ #
    # reboot: перезавантажити AI
    # ------------------------------------------------------------------ #
    if action == "reboot":
        task = user_tasks.pop(user_id, None)
        if task and not task.done():
            task.cancel()
        agent.reset_chat(user_id)
        user_context.pop(user_id, None)
        await message.answer("🔄 AI перезапущено. Контекст очищено!")
        return

    # ------------------------------------------------------------------ #
    # schedule_stop: зупинити таймер
    # ------------------------------------------------------------------ #
    if action == "schedule_stop":
        task = scheduled_tasks.pop(user_id, None)
        if task and not task.done():
            task.cancel()
        await delete_schedule(user_id)
        await message.answer("⏹ Таймер зупинено. Автоматичний пошук вимкнено.")
        return

    # ------------------------------------------------------------------ #
    # schedule_list: показати активні таймери
    # ------------------------------------------------------------------ #
    if action == "schedule_list":
        if user_id in scheduled_tasks and not scheduled_tasks[user_id].done():
            schedules = await get_all_schedules()
            info = next((s for s in schedules if s["user_id"] == user_id), None)
            if info:
                plats = " + ".join(
                    PLATFORM_LABELS.get(p, p)
                    for p in info["platform"].split(",") if p
                )
                unit = (f"{info['interval_minutes']} хв"
                        if info["interval_minutes"] < 60
                        else f"{info['interval_minutes'] // 60} год")
                await message.answer(
                    f"⏰ Активний таймер:\n"
                    f"Запит: «{info['query']}»\n"
                    f"Інтервал: кожні {unit}\n"
                    f"Платформи: {plats}\n\n"
                    f"Щоб зупинити — напишіть «зупини таймер»."
                )
            else:
                await message.answer("Активних таймерів немає.")
        else:
            await message.answer("Активних таймерів немає.")
        return

    # ------------------------------------------------------------------ #
    # schedule_set: встановити таймер
    # ------------------------------------------------------------------ #
    if action == "schedule_set":
        query = intent.get("query", "").strip()
        interval = intent.get("interval_minutes")
        detected = intent.get("platforms") or []
        platforms = detected if detected else settings["platforms"]

        if not query or not interval:
            await message.answer(
                "❓ Не зрозумів. Вкажіть товар і інтервал, наприклад:\n"
                "«Кожні 30 хвилин шукати корм для собак на Prom»"
            )
            return

        old = scheduled_tasks.pop(user_id, None)
        if old and not old.done():
            old.cancel()

        platforms_str = ",".join(platforms)
        await save_schedule(user_id, message.chat.id, interval, query, platforms_str)
        _start_schedule_task(user_id, message.chat.id, interval, query, platforms)

        label_p = " + ".join(PLATFORM_LABELS.get(p, p) for p in platforms)
        unit = f"{interval} хв" if interval < 60 else f"{interval // 60} год"
        await message.answer(
            f"⏰ Таймер встановлено!\n"
            f"Кожні {unit} шукатиму «{query}» на {label_p}.\n"
            f"Щоб зупинити — напишіть «зупини таймер»."
        )
        return

    # ------------------------------------------------------------------ #
    # search_web_manual: відкрити гугл і пропарсити розділ Shopping
    # ------------------------------------------------------------------ #
    if action == "search_web_manual":
        # AI сам виділяє запит, бренд, вагу тощо.
        query = await asyncio.to_thread(agent.extract_search_query, user_id, text)
        
        if not query:
            await message.answer("❓ AI не зміг виділити товар із вашого запиту. Спробуйте написати чіткіше.")
            return

        status_msg = await message.answer(f"🌐 Шукаю «{query}» в Google Shopping...\n⏳ Це займе до 5 хвилин.")
        
        from scraper_web import WebScraper
        web_scraper = WebScraper()
        
        # Функція для виконання пошуку та відправки результатів
        async def web_search_task():
            try:
                # 1. Збираємо "сирі" дані (текстові блоки + URL)
                raw_blocks_data = await asyncio.to_thread(web_scraper.open_google_manual, query)
                
                if not raw_blocks_data:
                    await status_msg.edit_text("😔 Не вдалося знайти товари або виникла помилка доступу.")
                    return

                await status_msg.edit_text(f"📦 Знайдено {len(raw_blocks_data)} карток. AI розпізнає ціни та магазини...")

                # 2. Витягуємо фільтри з контексту
                intent = await asyncio.to_thread(agent.classify_intent, user_id, text)
                filters = intent.get("filters") or {}

                # 3. AI парсить блоки паралельно у 2 потоки
                mid = len(raw_blocks_data) // 2
                half_a = raw_blocks_data[:mid] if mid else raw_blocks_data
                half_b = raw_blocks_data[mid:] if mid else []
                ai_tasks = [asyncio.to_thread(
                    agent.parse_raw_shopping_data,
                    user_id, [b["raw_text"] for b in half_a], query, filters,
                )]
                if half_b:
                    ai_tasks.append(asyncio.to_thread(
                        agent.parse_raw_shopping_data,
                        user_id, [b["raw_text"] for b in half_b], query, filters,
                    ))
                ai_results = await asyncio.gather(*ai_tasks)
                parsed_products = ai_results[0] + (ai_results[1] if len(ai_results) > 1 else [])
                raw_blocks_data = half_a + half_b  # порядок збережено для URL за індексом

                # Додаємо URL назад до розпарсених товарів (співставляємо за індексом)
                # Дедублікація: seller + name[:40] + price — всі три збіглись = дублікат
                final_filtered = []
                seen_keys: set[str] = set()
                for i, p in enumerate(parsed_products):
                    if i < len(raw_blocks_data) and p.get("match"):
                        seller = p.get("seller", "").strip().lower()
                        name = p.get("name", "").strip().lower()[:40]
                        price = p.get("price", "").strip()
                        dedup_key = f"{seller}|{name}|{price}"
                        if dedup_key in seen_keys:
                            continue
                        seen_keys.add(dedup_key)
                        p["url"] = raw_blocks_data[i]["url"]
                        p["platform"] = "google_shopping"
                        final_filtered.append(p)

                if not final_filtered:
                    await status_msg.edit_text("❌ AI не знайшов товарів, що відповідають вашому запиту.")
                    return

                # 4. Групуємо по смакам/підвидам
                final_filtered = await asyncio.to_thread(agent.group_products_by_subtype, user_id, final_filtered)

                # 5. Формуємо аналіз та Excel
                ai_analysis = await asyncio.to_thread(agent.analyze_prices, user_id, query, final_filtered, "Google Shopping")
                xlsx_bytes = await asyncio.to_thread(build_excel, query, "google_shopping", final_filtered, ai_analysis)
                
                filename = f"google_shopping_{query[:20].replace(' ', '_')}.xlsx"
                await status_msg.delete()
                await bot.send_document(
                    message.chat.id,
                    BufferedInputFile(xlsx_bytes, filename=filename),
                    caption=f'📊 Результати з Google Shopping для: "{query}"\n✅ Знайдено та розпізнано: {len(final_filtered)} тов.'
                )
            except Exception as e:
                logger.error("Web search task error: %s", e)
                await status_msg.edit_text("❌ Виникла помилка під час обробки даних AI.")

        # Запускаємо процес
        task = asyncio.create_task(web_search_task())
        user_tasks[user_id] = task
        return

    # ------------------------------------------------------------------ #
    # search: знайти товар
    # ------------------------------------------------------------------ #
    if action == "search":
        query = intent.get("query", "").strip()
        detected = intent.get("platforms") or []
        platforms = detected if detected else settings["platforms"]
        filters = intent.get("filters") or {}

        if not query:
            await message.answer("❓ Не вдалося визначити що шукати. Уточніть, будь ласка.")
            return

        # Формуємо підказку про активні фільтри
        filter_hints = []
        if filters.get("weight_kg") is not None:
            filter_hints.append(f"вага: {filters['weight_kg']} кг")
        if filters.get("price_min") is not None:
            filter_hints.append(f"від {filters['price_min']} грн")
        if filters.get("price_max") is not None:
            filter_hints.append(f"до {filters['price_max']} грн")
        if filters.get("brand"):
            filter_hints.append(f"бренд: {filters['brand']}")
        filter_str = f"\nФільтри: {', '.join(filter_hints)}" if filter_hints else ""

        label = " + ".join(PLATFORM_LABELS.get(p, p) for p in platforms)
        status_msg = await message.answer(
            f"🔎 Збираю «{query}» на {label}...{filter_str}\n⏳ Це може зайняти трохи часу."
        )

        # filter_intent передаємо лише якщо є реальні уточнення
        # (слова-квалфікатори АБО структуровані фільтри).
        # Інакше завжди вмикається СУВОРИЙ режим і AI відкидає половину товарів.
        has_qualifiers = (
            any(q in text.lower() for q in FILTER_QUALIFIERS)
            or bool(filters and any(v is not None for v in filters.values()))
        )
        filter_intent_param = text if has_qualifiers else ""

        async def search_task():
            await do_search(user_id, message.chat.id, query, platforms, status_msg,
                            filter_intent=filter_intent_param, filters=filters)

        task = asyncio.create_task(search_task())
        user_tasks[user_id] = task
        try:
            await task
        except asyncio.CancelledError:
            try:
                await status_msg.delete()
            except Exception:
                pass
            await message.answer("⛔ Пошук скасовано.")
        finally:
            user_tasks.pop(user_id, None)
        return

    # ------------------------------------------------------------------ #
    # chat: звичайна розмова
    # ------------------------------------------------------------------ #
    thinking = await message.answer("💭 Думаю...")
    reply = await asyncio.to_thread(agent.follow_up, user_id, text)
    await thinking.delete()
    await message.answer(reply)


@dp.message(F.photo)
async def handle_photo(message: Message) -> None:
    user_id = message.from_user.id
    settings = await get_user_settings(user_id)
    platforms = settings["platforms"]

    status_msg = await message.answer("🖼 Аналізую фото за допомогою AI...")

    photo = message.photo[-1]
    file = await bot.get_file(photo.file_id)
    file_bytes = await bot.download_file(file.file_path)
    image_bytes = file_bytes.read()

    caption = (message.caption or "").strip()

    product_name = await asyncio.to_thread(agent.identify_product_from_image, image_bytes)

    if not product_name:
        await status_msg.edit_text(
            "😔 Не вдалося визначити товар на фото. "
            "Спробуйте надіслати чіткіше фото або введіть назву вручну."
        )
        return

    # Витягуємо структуровані фільтри з підпису (якщо є)
    caption_filters: dict = {}
    if caption:
        await asyncio.to_thread(agent.add_user_message, user_id, caption)
        caption_intent = await asyncio.to_thread(agent.classify_intent, user_id, caption)
        caption_filters = caption_intent.get("filters") or {}

    label = " + ".join(PLATFORM_LABELS.get(p, p) for p in platforms)
    await status_msg.edit_text(
        f"✅ Визначено: *{product_name}*\n🔎 Шукаю на {label}...",
        parse_mode="Markdown",
    )
    await do_search(
        user_id, message.chat.id, product_name, platforms, status_msg,
        filter_intent=caption, filters=caption_filters,
    )


# ------------------------------------------------------------------ #
#  Entry point                                                         #
# ------------------------------------------------------------------ #

async def main() -> None:
    logger.info("Bot is starting...")
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
